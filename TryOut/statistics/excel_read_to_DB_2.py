# https://remusao.github.io/posts/few-tips-sqlite-perf.html

import sqlite3
import openpyxl
from openpyxl import load_workbook
from icecream import ic

# excel_file_name = r"C:\Users\kazak.ke\Documents\Задачи\Парсинг_параметризация\SRC\Статистика_20.xlsx"
# excel_file_name = r"F:\Kazak\GoogleDrive\1_KK\Job_CNAC\АИС\Статистика_21.xlsx"
excel_file_name = r"statistics.xlsx"

create_query = """
    CREATE TABLE IF NOT EXISTS tblRawStatistics ( 
        CASE_NUM            TEXT, 
        PRESSMARK           TEXT, 
        TITLE               TEXT, 
        UOM                 TEXT, 
        VOLUME              TEXT, 
        CARD_NUM            TEXT, 
        OBJ_NAME            TEXT, 
        ACTIVITY_TYPE       TEXT, 
        OBJ_TYPE            TEXT, 
        OBJ_GROUP           TEXT, 
        EXAMINATION_DATE    TEXT
);"""

insert_query = """
    INSERT INTO tblRawStatistics (
        CASE_NUM, PRESSMARK, TITLE, UOM, VOLUME, CARD_NUM, OBJ_NAME, 
        ACTIVITY_TYPE, OBJ_TYPE, OBJ_GROUP, EXAMINATION_DATE) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
"""

connection = sqlite3.connect('statistics_raw.sqlite')

with connection:
    connection.execute(create_query)

try:
    with connection:
        connection.executemany('INSERT INTO events VALUES (?, ?)', [
            (1, 'foo'),
            (2, 'bar'),
            (3, 'baz'),
            (1, 'foo'),
        ])
except (sqlite3.OperationalError, sqlite3.IntegrityError) as e:
    print('ошибка вставки:', e)


for row in connection.execute('SELECT * FROM tblRawStatistics'):
    print(row)

connection.close()






# PRESSMARK	TITLE	UOM	VOLUME	CARD_NUM	OBJ_NAME	ACTIVITY_TYPE	OBJ_TYPE	OBJ_GROUP	EXAMINATION_DATE



con.execute(create_query)

d1 = ('МГЭ/34877-1/5', '0.0-0-0', 'МАССА МУСОРА', 'т', '22.08', '77-10171/20-(0)-0', 'Капитальный ремонт многоквартирного дома. г. Москва, улица Строителей, д. 4, к. 4 (ЮЗАО, Гагаринский)', 'Капитальный ремонт', 'Производственного и непроизводственного назначения', 'Многоэтажный многоквартирный жилой дом', '2021-01-13 00:00:00')

wb = load_workbook(filename=excel_file_name, data_only=True)
sheets = wb.sheetnames
ic(sheets)

for sheet in sheets:
    ws = wb[sheet]
    ic(ws.title)
    ic(type(ws.rows))
    x = next(ws.rows)

    for row in ws.iter_rows(min_row=2):
        column_max = 10
        data = tuple([str(row[i].value) for i in range(column_max+1)])
        print(data)
        con.execute(insert_query, data)

i = con.execute("""PRAGMA table_info(tblRawStatistics);""")
print(i.fetchall())

