# https://remusao.github.io/posts/few-tips-sqlite-perf.html
# https://stackabuse.com/a-sqlite-tutorial-with-python/
# https://charlesleifer.com/blog/going-fast-with-sqlite-and-python/

import sqlite3
import openpyxl
from openpyxl import load_workbook
from datetime import datetime


from icecream import ic
import os

excel_file_name = r"C:\Users\kazak.ke\Documents\Задачи\Парсинг_параметризация\SRC\Статистика_20.xlsx"
# excel_file_name = r"F:\Kazak\GoogleDrive\1_KK\Job_CNAC\АИС\Статистика_21.xlsx"

DEFAULT_PATH = os.path.join(os.path.dirname(__file__), 'statistics_raw.sqlite3')
ic(DEFAULT_PATH)

# excel_file_name = r"statistics.xlsx"

create_query = """
    CREATE TABLE IF NOT EXISTS tblRawStatistics ( 
        CASE_NUM            TEXT, 
        PRESSMARK           TEXT, 
        TITLE               TEXT, 
        UOM                 TEXT, 
        VOLUME              TEXT, 
        CARD_NUM            TEXT, 
        OBJ_NAME            TEXT, 
        ACTIVITY_TYPE       TEXT, 
        OBJ_TYPE            TEXT, 
        OBJ_GROUP           TEXT, 
        EXAMINATION_DATE    TEXT
);"""

insert_query = """
    INSERT INTO tblRawStatistics (
        CASE_NUM, PRESSMARK, TITLE, UOM, VOLUME, CARD_NUM, OBJ_NAME, 
        ACTIVITY_TYPE, OBJ_TYPE, OBJ_GROUP, EXAMINATION_DATE) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
"""

data = ('МГЭ/34877-1/5', '0.0-0-0', 'МАССА МУСОРА', 'т', '22.08', '77-10171/20-(0)-0',
        'Капитальный ремонт многоквартирного дома. г. Москва, улица Строителей, д. 4, к. 4 (ЮЗАО, Гагаринский)',
        'Капитальный ремонт',
        'Производственного и непроизводственного назначения',
        'Многоэтажный многоквартирный жилой дом',
        '2021-01-13 00:00:00')

connection = sqlite3.connect(DEFAULT_PATH)
with connection:
    connection.execute(create_query)
    ic(connection.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall())
    # cursor = connection.execute(
    # """SELECT sql FROM sqlite_master WHERE type='table' AND name='tblRawStatistics'"""
    # ).fetchone()[0]
    # ic(cursor)
    for field in connection.execute("""PRAGMA table_info(tblRawStatistics);"""):
        ic(field)


def insert_db(in_connection: sqlite3.Connection, src_data: tuple):
    try:
        with in_connection as con:
            cursor = con.execute(insert_query, data)
            # ic(cursor.lastrowid)
            con.commit()
    except (sqlite3.OperationalError, sqlite3.IntegrityError) as err:
        in_connection.rollback()
        print('ошибка вставки:', err)


wb = load_workbook(filename=excel_file_name, data_only=True)
wb.iso_dates = False
sheets = wb.sheetnames
ic(sheets)
# from openpyxl.styles import Style
# s = openpyxl.Style(number_format=NumberFormat("dd-mm-yyyy"))
# ws["A2"].number_format = "0.00" # Display to 2dp.number_format


for sheet in sheets:
    ws = wb[sheet]
    ic(ws.title)
    # ic(type(ws.rows))
    for row in ws.iter_rows(min_row=2):
        column_max = 10
        data = [str(row[i].value) for i in range(column_max + 1)]
        data[10]=datetime.strptime(data[10], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
        data = tuple(data)
        # print(data)
        insert_db(connection, data)
    ic(connection.execute('SELECT COUNT(*) FROM tblRawStatistics').fetchall())

wb.close()

# for row in connection.execute('SELECT * FROM tblRawStatistics'):
#     print(row)
# ic(connection.execute('SELECT COUNT(*) FROM tblRawStatistics').fetchall())
connection.close()




